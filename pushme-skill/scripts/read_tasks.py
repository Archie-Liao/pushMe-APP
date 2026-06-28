"""读取用户的 Excel 任务文件，输出当前任务清单。"""
import openpyxl
import sys
import os
import io
from datetime import datetime

# 强制 UTF-8 输出，避免 Windows GBK 编码错误
# 注意：必须同时处理 stdout 和 stderr，且兼容管道模式（无 .buffer）
def _force_utf8():
    for attr in ('stdout', 'stderr'):
        stream = getattr(sys, attr)
        try:
            if hasattr(stream, 'buffer'):
                setattr(sys, attr, io.TextIOWrapper(stream.buffer, encoding='utf-8'))
        except (ValueError, AttributeError):
            pass  # 管道或无 buffer 时不强转
_force_utf8()

def read_task_board(filepath):
    """读取任务榜.xlsx，返回任务列表。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tasks = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [str(c).strip() if c is not None else '' for c in row]
            if any(v for v in vals):
                tasks.append({
                    'sheet': sn,
                    'type': vals[0] if len(vals) > 0 else '',
                    'title': vals[1] if len(vals) > 1 else '',
                    'detail': vals[2] if len(vals) > 2 else '',
                    'next_step': vals[3] if len(vals) > 3 else '',
                })
    return tasks

def read_archie_journal(filepath):
    """读取 Archie修仙录.xlsx，返回自我条例和机缘录。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
            vals = [str(c).replace('\n', ' | ').strip() if c is not None else '' for c in row]
            rows_data.append(vals)
        result[sn] = rows_data
    return result

if __name__ == '__main__':
    # 默认路径，可命令行覆盖
    if len(sys.argv) < 2:
        print('用法: python read_tasks.py <任务榜.xlsx路径> [Archie修仙录.xlsx路径]')
        print('示例: python read_tasks.py "C:\\Users\\...\\任务榜.xlsx" "C:\\Users\\...\\Archie修仙录.xlsx"')
        sys.exit(1)
    task_path = sys.argv[1]
    archie_path = sys.argv[2] if len(sys.argv) > 2 else None

    if os.path.exists(task_path):
        tasks = read_task_board(task_path)
        print(f'=== 任务榜 ({len(tasks)} entries) ===')
        for t in tasks:
            print(f"[{t['type']}] {t['title']}")
            if t['detail']:
                print(f"  详情: {t['detail'][:100]}")
            if t['next_step']:
                print(f"  下一步: {t['next_step'][:100]}")
            print()

    if archie_path and os.path.exists(archie_path):
        journal = read_archie_journal(archie_path)
        for sn, rows in journal.items():
            print(f'=== {sn} ({len(rows)} rows) ===')
            for r in rows[:5]:  # 只显示前 5 行作为摘要
                print(' | '.join(r[:3]))
            print(f'... 共 {len(rows)} 行')
            print()
